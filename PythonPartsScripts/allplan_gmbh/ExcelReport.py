""" Script forExcel Report
"""

from __future__ import annotations

from typing import Any

import NemAll_Python_AllplanSettings as AllplanSettings
import NemAll_Python_BaseElements as AllplanBaseEle
import NemAll_Python_Geometry as AllplanGeo
import NemAll_Python_IFW_ElementAdapter as AllplanEleAdapter
import NemAll_Python_IFW_Input as AllplanIFW
import NemAll_Python_Utility as AllplanUtil
import openpyxl
import openpyxl.styles as excel_style

from BaseInteractor import BaseInteractor
from BuildingElement import BuildingElement
from BuildingElementComposite import BuildingElementComposite
from BuildingElementControlProperties import BuildingElementControlProperties
from BuildingElementPaletteService import BuildingElementPaletteService
from ControlPropertiesUtil import ControlPropertiesUtil
from CreateElementResult import CreateElementResult
from DocumentManager import DocumentManager
from StringTableService import StringTableService
from Utils import LibraryBitmapPreview

print('Load ExcelReport.py')


def check_allplan_version(build_ele: BuildingElement,
                          version  : str) -> bool:
    """ Check the current Allplan version

    Args:
        build_ele: building element with the parameter properties
        version:   the current Allplan version

    Returns:
        True
    """

    # Support all versions
    del (build_ele, version)
    return True


def create_preview(build_ele: BuildingElement,
                   doc      : AllplanEleAdapter.DocumentAdapter) -> CreateElementResult:
    """ Creation of the element preview

    Args:
        build_ele: building element with the parameter properties
        doc:       document of the Allplan drawing files

    Returns:
        created elements for the preview
    """
    del (build_ele, doc)
    return CreateElementResult(LibraryBitmapPreview.create_library_bitmap_preview( \
                               AllplanSettings.AllplanPaths.GetPythonPartsEtcPath() +
                               r"Library\PythonParts\Wood_Construction.png"))

def create_interactor(coord_input              : AllplanIFW.CoordinateInput,
                      pyp_path                : str,
                      global_str_table_service: StringTableService,
                      build_ele_list           : list[BuildingElement],
                      build_ele_composite      : BuildingElementComposite,
                      control_props_list       : list[BuildingElementControlProperties],
                      modify_uuid_list        : list) -> object:
    """ Create the interactor

    Args:
        coord_input:               API object for the coordinate input, element selection, ... in the Allplan view
        pyp_path:                 path of the pyp file
        global_str_table_service: global string table service
        build_ele_list:            list with the building elements
        build_ele_composite:       building element composite with the building element constraints
        control_props_list:        control properties list
        modify_uuid_list:         list with the UUIDs of the modified elements

    Returns:
          Created interactor object
    """
    del (pyp_path, global_str_table_service, modify_uuid_list)
    return ExcelReport(coord_input, build_ele_list, build_ele_composite, control_props_list)


class ExcelReport(BaseInteractor):
    """ Definition of class Excel report
    """

    def __init__(self,
                 coord_input        : AllplanIFW.CoordinateInput,
                 build_ele_list     : list[BuildingElement],
                 build_ele_composite: BuildingElementComposite,
                 control_props_list : list[BuildingElementControlProperties]):
        """ Create the interactor

        Args:
            coord_input:         API object for the coordinate input, element selection, ... in the Allplan view
            build_ele_list:      list with the building elements
            build_ele_composite: building element composite with the building element constraints
            control_props_list:  control properties list
        """

        self.coord_input    = coord_input
        self.build_ele_list = build_ele_list
        self.build_ele      = build_ele_list[0]
        self.ctrl_prop_util = ControlPropertiesUtil(control_props_list, build_ele_list)

        self.palette_service = BuildingElementPaletteService(build_ele_list, build_ele_composite,
                                                             self.build_ele.script_name,
                                                             control_props_list, self.build_ele.pyp_file_path+"\\")

        self.palette_service.show_palette(self.build_ele.pyp_file_name)


        #----------------- set the filter and start the selection

        self.post_element_selection = None
        self.interactor_input_mode = None

        self.all_object_list = AllplanEleAdapter.BaseElementAdapterList()
        self.selected_elements = AllplanEleAdapter.BaseElementAdapterList()

        self.report_worksheet_path = ""
        """Path to the excel sheet for the report"""

        self.attribut_selection = [0]
        """Selection of the attributes for the parameter report"""

        self.object_ident = ""
        """the objects identity attribute"""

        self.qto_selection = [0]
        """Selection of the attributes for the qto report"""


    def start_element_selection(self):
        """ start the element selection
        """

        sel_setting = AllplanIFW.ElementSelectFilterSetting()

        self.post_element_selection = AllplanIFW.PostElementSelection()

        AllplanIFW.InputFunctionStarter.StartElementSelect("Objekte wählen!",
                                                           sel_setting, self.post_element_selection, True)

        self.interactor_input_mode = self.InteractorInputMode.ELEMENT_SELECTION

        #self.build_ele.IsInSelection.value = True

    def modify_element_property(self,
                                page : int,
                                name : str,
                                value: str):
        """ Modify property of element

        Args:
            page:  page index of the modified property
            name:  name of the modified property
            value: new value
        """
        palette_update_necessary = self.palette_service.modify_element_property(page, name, value)

        if name == "object_ident":
            self.update_group_list(self.build_ele.AttributeIDList.value)

        elif "read_tabel_path" in name:
            self.report_worksheet_path = self.build_ele.read_tabel_path.value

            report_template = openpyxl.load_workbook(
                self.report_worksheet_path,
                read_only  = True,
                data_only  = False,
                keep_links = False)

            sheet_list = report_template.sheetnames
            sheet_list.insert(0,"")
            sheets = "|".join(sheet_list)
            self.ctrl_prop_util.set_value_list("read_sheet_name",sheets)

        elif "save_tabel_path" in name:
            self.report_worksheet_path = self.build_ele.save_tabel_path.value

        if palette_update_necessary:
            self.palette_service.update_palette(-1, False)


    def set_active_palette_page_index(self,
                                      active_page_index: int):

        """ Handles the event of changing the page in the property palette and a dialog
        (e.g. open file dialog) being closed.

        Switches to PLACE mode directly after selecting a VS-PythonPart

        Args:
            active_page_index: index of the active page, starting from 0
        """

        if self.qto_selection != self.build_ele.QTOAttributeList.value:
            self.qto_selection = self.build_ele.QTOAttributeList.value

            print ("Select the qto attributes")


        attrib_group_IDs = set(self.build_ele.AttributeIDList.value)

        if self.attribut_selection != attrib_group_IDs:
            print("the choosen attributes list has changed")
            self.update_group_list(attrib_group_IDs)

        del active_page_index

    def update_group_list(self,
                          attrib_group_IDs  :set[int]):
        

        """ updates the list of attributes in the pulldown from which the
            grouping attribut for the report can be choosen

            update is executed each time when a new attribut is choosen
            or an already choosen one is removed

        Args:
            attrib_group_IDs: list of the attribut IDs choosen as report parameters
        """

        self.attribut_selection = attrib_group_IDs
        doc = DocumentManager.get_instance().document
        possible_group_attribs = []

        (local_str_table, global_str_table) = self.build_ele.get_string_tables()
        free_ident_attrib = local_str_table.get_string("2002", "Choose...")

        if self.build_ele.object_ident.value == free_ident_attrib:
            object_ident_name = AllplanBaseEle.AttributeService.GetAttributeName(doc,self.build_ele.object_ident_ID.value)
        else:
            object_ident_name = self.build_ele.object_ident.value

        possible_group_attribs.append(object_ident_name)


        for single_ID in self.attribut_selection:
            single_name = AllplanBaseEle.AttributeService.GetAttributeName(doc, single_ID)
            possible_group_attribs.append(single_name)
            possible_group_attribs.insert(0, "")
            if "???" in possible_group_attribs:
                possible_group_attribs.remove("???")
            group_name = "|".join(possible_group_attribs)

        self.ctrl_prop_util.set_value_list("group_attrib",group_name)
        self.palette_service.update_palette(-1, False)

        del global_str_table

    def on_control_event(self,
                         event_id: int):
        """ Handles on control event

        Args:
            event_id: event id of the clicked button control

        """
        (local_str_table, global_str_table) = self.build_ele.get_string_tables()
        export_message = local_str_table.get_string("2001", "Excel report is created!")

        if event_id == 1002:
            self.start_element_selection()

        if event_id == 1005:
            ret = AllplanUtil.ShowMessageBox (export_message, AllplanUtil.MB_OK)
            self.create_excel_tabel(self.build_ele)

            print("Return value = " , ret, export_message, "= ", ret & AllplanUtil.MB_OKCANCEL)

        del global_str_table

    def on_cancel_function(self) -> bool:
        """ Handles the cancel function event (e.g. by ESC, ...)

        Returns:
            True/False for success.
        """
        if self.interactor_input_mode == self.InteractorInputMode.ELEMENT_SELECTION and self.post_element_selection:
            return False


        self.palette_service.close_palette()

        return True

    def on_preview_draw(self):
        """ Handles the preview draw event
        """


    def on_mouse_leave(self):
        """ Handles the mouse leave event
        """
        self.on_preview_draw()

    def on_value_input_control_enter(self) -> bool:
        """ Handles the enter inside the value input control event

        Returns:
            True/False for success.
        """

        return True

    def process_mouse_msg(self,
                          mouse_msg: int,
                          pnt      : AllplanGeo.Point2D,
                          msg_info : Any) -> bool:
        """ Process the mouse message event

        Args:
            mouse_msg: mouse message ID
            pnt:       input point in Allplan view coordinates
            msg_info:  additional mouse message info

        Returns:
            True/False for success.
        """
        if self.interactor_input_mode == self.InteractorInputMode.ELEMENT_SELECTION and self.post_element_selection:
            self.selected_elements = self.post_element_selection.GetSelectedElements(self.coord_input.GetInputViewDocument())

            for single_element in self.selected_elements:
                self.all_object_list.append(single_element)

            self.post_element_selection = None

            self.palette_service.update_palette(-1, True)

            return True


        if self.coord_input.IsMouseMove(mouse_msg):
            return True

        if self.coord_input.IsMouseMove(mouse_msg):
            return True

        self.palette_service.update_palette(-1, True)

        return True

    def create_excel_tabel (self,
                            build_ele   : BuildingElement) -> bool:
        """ creating the Excel table with the selected attributes and
            parameters either in an already existing file or in a
            new one. the values for the parameters of the selected object
            are writen row by row - one row for each object 

        Args:
            build_ele: building element with the parameter properties

        Returns:
            True/False for success.
        """

        object_list = self.all_object_list
        selected_object_list = AllplanEleAdapter.BaseElementAdapterList()

        (local_str_table, global_str_table) = build_ele.get_string_tables()
        free_ident = local_str_table.get_string("2002", "Choose...")


        doc = DocumentManager.get_instance().document

        #----------------- Extract palette parameter values
        #----------------- For both kinds of report

        kind_of_report = build_ele.Report_kind.value

        #-------- create or write
        excel_source = build_ele.tabel_handling.value
        if excel_source == "use":
            excel_path = build_ele.read_tabel_path.value
            excel_sheet = build_ele.read_sheet_name.value
        else:
            excel_path = build_ele.save_tabel_path.value
            excel_sheet = "Report"

        excel_row = build_ele.start_row.value
        excel_column = build_ele.start_column.value

        #----------------- For parameter report

        object_identity = build_ele.object_ident.value
        object_ident_ID = 498
        if object_identity == "Allright_ID":
            object_ident_ID = 10
        if object_identity == "IFC_ID":
            object_ident_ID = 683
        if object_identity == free_ident:
            object_ident_ID = build_ele.object_ident_ID.value


        selected_attributes = build_ele.AttributeIDList.value
        grouping_attribute = build_ele.group_attrib.value

        #----------------- For qto of report

        choosen_qto_attributes = build_ele.QTOAttributeList.value
        param_attribut = build_ele.qto_ident_attrib.value



        #----------------- Open report template

        if excel_source == "use":
            report_file = openpyxl.load_workbook(excel_path, read_only=False, data_only=False, keep_links=False)

        else:
            report_file = openpyxl.Workbook()
            report_file.create_sheet(excel_sheet)

        if "Sheet" in report_file.sheetnames:
            report_file.remove(report_file["Sheet"])

        report_sheet = report_file[excel_sheet]


        #----------------- find template headers

        cell_header_font = excel_style.Font(bold = True, size = 14)
        cell_header_color = excel_style.PatternFill(fill_type = "solid", start_color = "BFBFBF")
        cell_header_border = excel_style.Side(border_style = "thick", color = "000000")
        cell_header_frame = excel_style.Border(left = cell_header_border, right = cell_header_border
                                               , bottom = cell_header_border, top = cell_header_border)

        cell_font = excel_style.Font(bold = False, size = 12)
        cell_color = excel_style.PatternFill(fill_type = "solid", start_color = "F8F8F8")

        #----------------- In case of parameter report
        if kind_of_report == "param":

            column_header_list = []
            column_header_list.append("Objekt")

            for attrib_ID in selected_attributes:
                if attrib_ID != 0 and attrib_ID is not None:
                    attrib_name = AllplanBaseEle.AttributeService.GetAttributeName(doc, attrib_ID)
                    column_header_list.append(str(attrib_name))

            group_ID = AllplanBaseEle.AttributeService.GetAttributeID(doc, grouping_attribute)

        #----------------- In case of qto report
        if kind_of_report == "qto":

            column_header_list = []
            qto_attrib_name = AllplanBaseEle.AttributeService.GetAttributeName(doc, param_attribut)
            column_header_list.append(str(qto_attrib_name))

            for qto_attrib_ID in choosen_qto_attributes:
                if qto_attrib_ID != 0 and qto_attrib_ID is not None:
                    param_name = AllplanBaseEle.AttributeService.GetAttributeName(doc, qto_attrib_ID)
                    column_header_list.append(str(param_name))

        #----------------- create template headers
        start_column = excel_column

        for column_header in column_header_list:
            excel_cell = report_sheet.cell(row = excel_row, column = start_column)
            excel_cell.value = column_header
            excel_cell.font = cell_header_font
            excel_cell.fill = cell_header_color
            excel_cell.border = cell_header_frame

            start_column += 1


        #----------------- Filter relevant objects

        for single_object in object_list:
            if single_object.IsInActiveLayer() is True:
                selected_object_list.append(single_object)
            else:
                continue

        #----------------- In case of parameter report
        if kind_of_report == "param":

            relevant_object_list = []

            for selected_object in selected_object_list:
                checking_attributes = AllplanBaseEle.ElementsAttributeService.GetAttributes(selected_object)
                for attrib_pair in checking_attributes:
                    if attrib_pair[0] == object_ident_ID:
                        relevant_object_list.append(selected_object)

            #----------------- Get objects attributes

            objects_attrib_list = []


            for relevant_object in relevant_object_list:
                objects_attrib_dict = {}
                grouping_value = "-"
                objects_attributes = AllplanBaseEle.ElementsAttributeService.GetAttributes(relevant_object)

                for attrib_pair in objects_attributes:
                    if attrib_pair[0] == object_ident_ID:
                        object_ident_value = attrib_pair[1]
                        if object_ident_value is None:
                            object_ident_value = "allgemeines Objekt"

                    if attrib_pair[0] == group_ID:
                        grouping_value = attrib_pair[1]

                    if attrib_pair[0] in selected_attributes:
                        objects_attrib_dict[attrib_pair[0]] = attrib_pair[1]

                objects_attrib_list.append((object_ident_value, objects_attrib_dict, grouping_value))


            #----------------- Group the selected objects

            existing_grouping_values = []

            for report_object in objects_attrib_list:
                objects_grouping_value = report_object[2]
                if objects_grouping_value not in existing_grouping_values:
                    existing_grouping_values.append(objects_grouping_value)

            objects_ordered_list = []
            object_random_list = []

            for existing_value in existing_grouping_values:
                checking_value = existing_value

                if existing_value is None or existing_value == 0:
                    checking_value = "-"

                for single_object in objects_attrib_list:
                    object_value = single_object[2]
                    if object_value == checking_value and checking_value != "-":
                        objects_ordered_list.append(single_object)
                    if checking_value == "-":
                        object_random_list.append(single_object)

            for random_object in object_random_list:
                objects_ordered_list.append(random_object)


            #----------------- Write attributes to Excel

            content_row = excel_row + 1

            for object_parameters in objects_ordered_list:
                objects_ID_name = object_parameters[0]
                parameter_dict = object_parameters[1]
                name_cell = report_sheet.cell(row = content_row, column = excel_column)
                name_cell.value = objects_ID_name
                name_cell.font = cell_font
                name_cell.fill = cell_color

                content_column = excel_column + 1


                for single_parameter in selected_attributes:
                    if single_parameter != 0 and single_parameter is not None:

                        if single_parameter in parameter_dict.keys():
                            param_value = parameter_dict.get(single_parameter)

                            attrib_type = isinstance(param_value, float)

                            if attrib_type == True:
                                param_value = f"{((round(param_value,3)*1000)/1000):.3f}"

                            if param_value == 0 or param_value is None:
                                param_value = "-"
                        else:
                            param_value = "-"

                        param_cell = report_sheet.cell(row = content_row, column = content_column)
                        param_cell.value = param_value
                        param_cell.font = cell_font
                        param_cell.fill = cell_color

                    content_column += 1


                content_row += 1

        #----------------- In case of qto report
        if kind_of_report == "qto":

            qto_object_list = []
            param_values_list = []
            param_quantity_list = []

            for selected_qto_object in selected_object_list:
                qto_attributes = AllplanBaseEle.ElementsAttributeService.GetAttributes(selected_qto_object)
                for param_pair in qto_attributes:
                    if param_pair[0] == param_attribut:
                        objects_param_value = param_pair[1]
                        if objects_param_value is not None and objects_param_value != 0 and objects_param_value != "-":
                            qto_object_list.append((objects_param_value, selected_qto_object))
                            param_values_list.append(objects_param_value)

            possible_param_values = set(param_values_list)

            for source_object in qto_object_list:
                source_attributes = AllplanBaseEle.ElementsAttributeService.GetAttributes(source_object[1])
                for source_attrib_pair in source_attributes:
                    if source_attrib_pair[0] in choosen_qto_attributes:
                        source_param = source_attrib_pair[0]
                        source_quantity = source_attrib_pair[1]
                        param_quantity_list.append((source_object[0], source_param, source_quantity))

                if 215 in choosen_qto_attributes:
                    source_count_param = 215
                    source_count_quantity = 1
                    param_quantity_list.append((source_object[0], source_count_param, source_count_quantity))

            param_qantity_dict_list = []

            for single_possible_value in possible_param_values:
                possible_value_dict = {}

                for param_quantity in param_quantity_list:
                    if param_quantity[0] == single_possible_value:
                        relevant_param = param_quantity[1]
                        if relevant_param not in possible_value_dict:
                            new_quantity = param_quantity[2]
                            possible_value_dict[relevant_param] = new_quantity
                        else:
                            existing_quantity = possible_value_dict.get(relevant_param)
                            new_quantity = existing_quantity + param_quantity[2]
                            possible_value_dict[relevant_param] = new_quantity

                param_qantity_dict_list.append((single_possible_value, possible_value_dict))

            #----------------- Write attributes to Excel

            content_row = excel_row + 1

            for named_param_quantity in param_qantity_dict_list:
                quantity_name = named_param_quantity[0]
                quantity_dict = named_param_quantity[1]
                name_cell = report_sheet.cell(row = content_row, column = excel_column)
                name_cell.value = quantity_name
                name_cell.font = cell_font
                name_cell.fill = cell_color

                content_column = excel_column + 1


                for single_quantity in choosen_qto_attributes:
                    if single_quantity != 0 and single_quantity is not None:

                        if single_quantity in quantity_dict.keys():
                            quantity_value = quantity_dict.get(single_quantity)

                            q_attrib_type = isinstance(quantity_value, float)

                            if q_attrib_type is True:
                                quantity_value = f"{((round(quantity_value,3)*1000)/1000):.3f}"

                            if quantity_value == 0 or quantity_value is None:
                                quantity_value = "-"
                        else:
                            quantity_value = "-"

                        param_cell = report_sheet.cell(row = content_row, column = content_column)
                        param_cell.value = quantity_value
                        param_cell.font = cell_font
                        param_cell.fill = cell_color

                    content_column += 1


                content_row += 1

        report_file.save(excel_path)
        report_file.close()
        del global_str_table
        return True
